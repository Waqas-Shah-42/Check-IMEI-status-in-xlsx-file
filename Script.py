import os
import Drs

from openpyxl import load_workbook

my_workbook = load_workbook("temp.xlsx")
worksheet = my_workbook.active

print('value of cell ', worksheet.cell(row=2, column=6).value, "\n\n")
print("rows ", worksheet.max_row)

'''
os.chdir(os.path.dirname(os.path.realpath(__file__)))

wb_read=xlrd.open_workbook("temp.xlsx")
data_read=wb_read.sheet_by_index(0)

wb_write = xlwt.open_workbook("temp.xlsx")
data_write=xlwt.sheet_by_index(0)
print(data_read.cell_value(1,5))

print("number of rows ",data_read.nrows)
'''

DRS = Drs.Drs()
status = "none initial"

no_of_rows = worksheet.max_row
for i in range(2, no_of_rows+1):
    status_temp = True
    try:
        status_temp = 'Approved'not in worksheet.cell(row=i, column=13).value or 'notApproved' in worksheet.cell(row=i, column=13).value
    except:
        status_temp = True
    finally:
        pass

    if(status_temp):
        try:
            imei = worksheet.cell(row=i, column=6).value
            status = DRS.IMEIStatus(imei)
            worksheet.cell(row=i, column=13).value = status.replace(" ", "")
            print(i, " ", imei, " ", status.replace(" ", ""), "\n")
        except:
            pass
        finally:
            pass

        try:
            device_detail = DRS.getDeviceDetails(imei)
            worksheet.cell(row=i, column=9).value = device_detail[1]
        except:
            pass
        finally:
            pass

    else:
        pass
        # print(i, ' Already approved')

        # try:
        #     imei = worksheet.cell(row=i, column=6).value
        #     device_detail = DRS.getDeviceDetails(imei)
        #     print('detail ',device_detail, '1st ', device_detail[0],'2nd ', device_detail[1])
        #     worksheet.cell(row=i, column=9).value = device_detail[1]
        # except:
        #     pass
        # finally:
        #     pass

my_workbook.save('balances.xlsx')
os.system('python "luhn check.py"')
