import os
# import xlrd
# import xlwt

from openpyxl import load_workbook
# import openpyxl

os.chdir(os.path.dirname(os.path.realpath(__file__)))


def luhn(imei):
    imei = str(imei)
    temp = ''
    # ensuring that no unwanted charcters are included in imei
    for char in imei:
        if char.isalnum():
            temp = temp + char

    imei = temp

    if(len(imei) > 15):
        imei = imei[0:15]

    # creating an array of imei digits
    if(len(imei) == 15):
        imei_arr = [0 for i in range(15)]
        for i in range(len(imei)):
            imei_arr[i] = int(imei[i])

        # iterating over digits where multiplication has to happen
        for i in range(1, len(imei), 2):
            temp_int = int(imei_arr[i])*2
            if temp_int > 9:
                # addition for luhns algorithim
                temp_int = temp_int//10 + temp_int % 10

            imei_arr[i] = temp_int

        # checking is sum of array is mod of 10
        if(sum(imei_arr) % 10 == 0):
            # print('passed luhns algorithim')
            return "passed"

        else:
            # print('failed luhns algorithim')
            return "failed"

    else:
        # luhns algorithim can only be carried out on IMEIs of size 15
        return 'not 15 length'




my_workbook = load_workbook("balances.xlsx")
worksheet = my_workbook.active

print('value of cell ', worksheet.cell(row=2, column=6).value, "\n\n")
print("rows ", worksheet.max_row)

worksheet.cell(row=3, column=16).value = "lolz"

no_of_rows = worksheet.max_row
for i in range(2, no_of_rows+1):
    imei1 = worksheet.cell(row=i, column=5).value
    imei2 = worksheet.cell(row=i, column=6).value

    worksheet.cell(row=i, column=14).value = luhn(imei1)
    worksheet.cell(row=i, column=15).value = luhn(imei2)

    # status = DRS.IMEIStatus(imei)
    # worksheet.cell(row=i, column=13).value = status
    # print(i, " ", imei1, " ", imei2, "\n")

my_workbook.save('temp.xlsx')
